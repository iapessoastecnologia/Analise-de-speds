import tkinter as tk
from tkinter import filedialog, messagebox
from ttkbootstrap import Style
from ttkbootstrap.widgets import Button
from PIL import Image, ImageTk, ImageSequence
from collections import defaultdict
import os
import re
import threading
import pandas
from pathlib import Path
import json
import chardet
import glob
import time
import shutil

# ========== CONFIGURAÇÃO INICIAL ==========
os.makedirs("noAI", exist_ok=True)
os.makedirs("resultado", exist_ok=True)

# ========== VARIÁVEIS GLOBAIS ==========
arquivo_selecionado = None
arquivos_selecionados = []  # NOVO: lista de arquivos selecionados
pasta_selecionada = None    # NOVO: pasta selecionada
processando = False
frames = []
gif_label = None
tempo_inicio = None
tempo_fim = None
temporizador_rodando = False
label_temporizador = None

# ========== FUNÇÕES ==========

def centralizar_janela(app, largura, altura):
    app.update_idletasks()
    largura_tela = app.winfo_screenwidth()
    altura_tela = app.winfo_screenheight()
    x = (largura_tela // 2) - (largura // 2)
    y = (altura_tela // 2) - (altura // 2)
    app.geometry(f"{largura}x{altura}+{x}+{y}")

def selecionar_arquivos():
    global arquivos_selecionados, arquivo_selecionado
    caminhos = filedialog.askopenfilenames(
        title="Selecione um ou mais arquivos Sped",
        filetypes=[("Arquivos TXT", "*.txt")]
    )
    if caminhos:
        arquivos_selecionados = list(caminhos)
        arquivo_selecionado = arquivos_selecionados[0] if arquivos_selecionados else None
        label_arquivo.config(text=f"Selecionados: {len(arquivos_selecionados)} arquivo(s)")
        botao_executar_analise.config(state="normal")
    else:
        arquivos_selecionados = []
        arquivo_selecionado = None
        label_arquivo.config(text="Nenhum arquivo selecionado.")
        botao_executar_analise.config(state="disabled")

def selecionar_pasta():
    global arquivos_selecionados, pasta_selecionada, arquivo_selecionado
    pasta = filedialog.askdirectory(title="Selecione a pasta com arquivos Sped")
    if pasta:
        arquivos = [os.path.join(pasta, f) for f in os.listdir(pasta) if f.lower().endswith('.txt')]
        arquivos_selecionados = arquivos
        pasta_selecionada = pasta
        arquivo_selecionado = arquivos_selecionados[0] if arquivos_selecionados else None
        label_arquivo.config(text=f"Selecionados: {len(arquivos_selecionados)} arquivo(s) da pasta")
        botao_executar_analise.config(state="normal" if arquivos_selecionados else "disabled")

def exportar_resultado():
    # Novo: Exporta os relatórios da pasta resultado para uma pasta escolhida
    pasta_destino = filedialog.askdirectory(title="Selecione a pasta de destino para exportar os relatórios")
    if not pasta_destino:
        return
    pasta_origem = os.path.abspath('resultado')
    try:
        for item in os.listdir(pasta_origem):
            origem = os.path.join(pasta_origem, item)
            destino = os.path.join(pasta_destino, item)
            if os.path.isdir(origem):
                if os.path.exists(destino):
                    shutil.rmtree(destino)
                shutil.copytree(origem, destino)
            else:
                shutil.copy2(origem, destino)
        messagebox.showinfo("Exportação Concluída", f"Relatórios exportados para: {pasta_destino}")
    except Exception as e:
        messagebox.showerror("Erro na Exportação", str(e))

def limpar_arquivos_finais():
    # Remove arquivos das pastas 'jsons' e 'noAI', mas não as pastas
    for pasta in ['jsons', 'noAI']:
        caminho = Path(pasta)
        if caminho.exists() and caminho.is_dir():
            for arquivo in caminho.iterdir():
                if arquivo.is_file():
                    try:
                        arquivo.unlink()
                    except Exception as e:
                        print(f"Erro ao apagar {arquivo}: {e}")
    # Após remover arquivos, remove as pastas inteiras se existirem
    for pasta in ['jsons', 'noAI']:
        caminho = Path(pasta)
        if caminho.exists() and caminho.is_dir():
            try:
                caminho.rmdir()
            except OSError:
                # Se não estiver vazia, remove tudo
                shutil.rmtree(str(caminho), ignore_errors=True)

def atualizar_temporizador():
    global tempo_inicio, temporizador_rodando
    if temporizador_rodando and tempo_inicio is not None:
        tempo_decorrido = int(time.time() - tempo_inicio)
        minutos = tempo_decorrido // 60
        segundos = tempo_decorrido % 60
        label_temporizador.config(text=f"Tempo: {minutos:02d}:{segundos:02d}")
        label_temporizador.after(1000, atualizar_temporizador)
    elif tempo_inicio is not None and tempo_fim is not None:
        tempo_total = int(tempo_fim - tempo_inicio)
        minutos = tempo_total // 60
        segundos = tempo_total % 60
        label_temporizador.config(text=f"Tempo: {minutos:02d}:{segundos:02d}")
    else:
        label_temporizador.config(text="Tempo: 00:00")

def iniciar_analise():
    global processando, tempo_inicio, tempo_fim, temporizador_rodando
    if not arquivos_selecionados:
        messagebox.showwarning("Aviso", "Nenhum arquivo selecionado.")
        return
    processando = True
    tempo_inicio = time.time()
    temporizador_rodando = True
    atualizar_temporizador()
    mostrar_animacao()
    botao_executar_analise.config(state="disabled")
    botao_exportar.config(state="disabled")
    threading.Thread(target=processar_arquivos_thread).start()

def processar_arquivos_thread():
    global processando, tempo_fim, temporizador_rodando
    try:
        filtro(arquivos_selecionados)
        organizar(arquivos_selecionados)
        verificar(arquivos_selecionados)
        messagebox.showinfo(
            "Sucesso",
            "Análise concluída! Os relatórios estão na pasta 'resultado'.\nVocê pode exportar novamente se desejar."
        )
        # Abrir a pasta de resultados automaticamente (opcional)
        try:
            pasta_resultado = os.path.abspath('resultado')
            if os.name == 'nt':
                os.startfile(pasta_resultado)
            elif os.name == 'posix':
                import subprocess
                subprocess.Popen(['xdg-open', pasta_resultado])
        except Exception as e:
            print(f"Não foi possível abrir a pasta de resultados: {e}")
    except Exception as e:
        messagebox.showerror("Erro", str(e))
    finally:
        processando = False
        tempo_fim = time.time()
        temporizador_rodando = False
        atualizar_temporizador()
        botao_exportar.config(state="normal")
        gif_label.config(image="")
        gif_label.image = None
        limpar_arquivos_finais()
        botao_executar_analise.config(state="normal")

# Animação de Carregamento
def mostrar_animacao():
    global frames, gif_label, processando
    try:
        caminho_gif = os.path.join('assets', 'animar.gif')
        if not os.path.exists(caminho_gif):
            print(f"[ERRO] GIF não encontrado em: {caminho_gif}")
            gif_label.config(image="")
            gif_label.image = None
            return
        gif = Image.open(caminho_gif)
        frames = [ImageTk.PhotoImage(frame.copy().resize((32, 32)).convert("RGBA")) for frame in ImageSequence.Iterator(gif)]
        if not frames:
            print("[ERRO] Nenhum frame carregado do GIF!")
            gif_label.config(image="")
            gif_label.image = None
            return
        def atualizar(indice):
            if not processando:
                gif_label.config(image="")
                gif_label.image = None
                return
            frame = frames[indice % len(frames)]
            gif_label.config(image=frame)
            gif_label.image = frame
            gif_label.update_idletasks()
            app.after(50, atualizar, indice + 1)
        gif_label.config(image=frames[0])
        gif_label.image = frames[0]
        gif_label.update_idletasks()
        atualizar(0)
        print("[DEBUG] GIF de carregamento exibido.")
    except Exception as e:
        print("Erro ao carregar animação:", e)
        gif_label.config(image="")
        gif_label.image = None

def sanitize_filename(name):
    # Remove caracteres inválidos para nomes de pastas/arquivos no Windows
    return re.sub(r'[<>:"/\\|?*]', '_', name)

# Filtra os arquivos SPED.txt para Deixar apenas dados relevantes[C100,C150,200 e Etc]
def filtro(arquivos):
    def analisar_arquivo(nome_arquivo):
        print(f"Analisando arquivo: {nome_arquivo}")
        if not os.path.exists(nome_arquivo):
            print(f"❌ Arquivo '{nome_arquivo}' não encontrado!")
            return False
        tamanho = os.path.getsize(nome_arquivo)
        print(f"📁 Tamanho do arquivo: {tamanho:,} bytes")
        with open(nome_arquivo, 'rb') as f:
            primeiros_bytes = f.read(1000)
        chars_binarios = sum(1 for b in primeiros_bytes if b < 32 and b not in [9, 10, 13])
        print(f"🔍 Caracteres de controle encontrados nos primeiros 1000 bytes: {chars_binarios}")
        texto_inicial = primeiros_bytes.decode('utf-8', errors='ignore')
        if '-----BEGIN' in texto_inicial or 'CERTIFICATE' in texto_inicial:
            print("🔐 Possível certificado digital detectado")
        return True

    def tentar_ler_arquivo(nome_arquivo):
        print("\n" + "="*60)
        print("🔄 INICIANDO LEITURA DO ARQUIVO")
        print("="*60)
        try:
            print("📊 Estratégia 1: Detectando encoding...")
            with open(nome_arquivo, 'rb') as f:
                raw_data = f.read()
                resultado = chardet.detect(raw_data)
                encoding = resultado['encoding']
                confianca = resultado['confidence']
                print(f"   ✅ Encoding detectado: {encoding} (confiança: {confianca:.2%})")
                if encoding and confianca > 0.7:
                    conteudo = raw_data.decode(encoding, errors='replace')
                    print(f"   ✅ Leitura bem-sucedida com {encoding}")
                    return conteudo, encoding
        except Exception as e:
            print(f"   ❌ Erro com chardet: {e}")
        
        print("\n📋 Estratégia 2: Testando encodings comuns...")
        encodings_para_tentar = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'windows-1252', 'cp850', 'ascii']
        for encoding in encodings_para_tentar:
            try:
                with open(nome_arquivo, 'r', encoding=encoding, errors='replace') as f:
                    conteudo = f.read()
                    print(f"   ✅ Sucesso com encoding: {encoding}")
                    return conteudo, encoding
            except Exception as e:
                print(f"   ❌ Falhou com {encoding}: {e}")
                continue

        print("\n🧹 Estratégia 3: Limpeza agressiva...")
        try:
            with open(nome_arquivo, 'rb') as f:
                raw_data = f.read()
                conteudo = raw_data.decode('utf-8', errors='ignore')
                conteudo = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', conteudo)
                conteudo = re.sub(r'[^\x20-\x7E\r\n\t\u00A0-\uFFFF]', '', conteudo)
                print("   ✅ Arquivo lido com limpeza agressiva")
                return conteudo, "utf-8-limpo"
        except Exception as e:
            print(f"   ❌ Erro na limpeza agressiva: {e}")
        print("   ❌ Todas as estratégias falharam")
        return None, None

    def limpar_certificados_e_assinaturas(conteudo):
        print("\n🔐 Removendo certificados digitais...")
        padroes_cert = [
            r'-----BEGIN CERTIFICATE-----.*?-----END CERTIFICATE-----',
            r'-----BEGIN.*?CERTIFICATE.*?-----.*?-----END.*?CERTIFICATE.*?-----',
            r'-----BEGIN[^-]*-----.*?-----END[^-]*-----',
            r'<ds:Signature.*?</ds:Signature>',
            r'<Signature.*?</Signature>',
            r'BEGIN PKCS7.*?END PKCS7',
            r'MII[A-Za-z0-9+/=]{100,}',
        ]
        conteudo_original = conteudo
        certificados_removidos = 0
        for padrao in padroes_cert:
            matches = re.findall(padrao, conteudo, flags=re.DOTALL | re.IGNORECASE)
            if matches:
                certificados_removidos += len(matches)
                conteudo = re.sub(padrao, '[CERTIFICADO/ASSINATURA REMOVIDA]', conteudo, flags=re.DOTALL | re.IGNORECASE)
        if certificados_removidos > 0:
            print(f"   ✅ {certificados_removidos} certificado(s)/assinatura(s) removido(s)")
            reducao = len(conteudo_original) - len(conteudo)
            print(f"   📉 Redução no tamanho: {reducao:,} caracteres")
        else:
            print("   ℹ️  Nenhum certificado/assinatura encontrado")
        return conteudo

    def filtrar_linhas_especificas(conteudo, codigos_desejados):
        print(f"\n🔍 FILTRANDO LINHAS ESPECÍFICAS")
        print(f"   📋 Códigos desejados: {', '.join(codigos_desejados)}")
        linhas = conteudo.split('\n')
        linhas_filtradas = []
        contadores = {codigo: 0 for codigo in codigos_desejados}
        total_linhas_originais = len(linhas)
        for linha in linhas:
            linha_limpa = linha.strip()
            if linha_limpa:
                for codigo in codigos_desejados:
                    if linha_limpa.startswith(codigo):
                        linhas_filtradas.append(linha)
                        contadores[codigo] += 1
                        break
        print(f"   📊 RESULTADO DO FILTRO:")
        print(f"      📝 Linhas originais: {total_linhas_originais:,}")
        print(f"      ✅ Linhas filtradas: {len(linhas_filtradas):,}")
        for codigo, count in contadores.items():
            msg = f"      🏷️  {codigo}: {count:,} linhas" if count > 0 else f"      ⚠️  {codigo}: 0 linhas (não encontrado)"
            print(msg)
        percentual = (len(linhas_filtradas) / total_linhas_originais) * 100 if total_linhas_originais > 0 else 0
        print(f"      📈 Percentual mantido: {percentual:.1f}%")
        return '\n'.join(linhas_filtradas)

    def salvar_arquivo_limpo(conteudo, nome_original, sufixo="_FILTRADO"):
        nome_limpo = nome_original.replace('.txt', f'{sufixo}.txt')
        try:
            with open(nome_limpo, 'w', encoding='utf-8', errors='replace') as f:
                f.write(conteudo)
            print(f"\n💾 Arquivo salvo: '{nome_limpo}'")
            print(f"   📏 Tamanho do conteúdo: {len(conteudo):,} caracteres")
            return nome_limpo
        except Exception as e:
            print(f"   ❌ Erro ao salvar arquivo: {e}")
            return None

    def mostrar_preview(conteudo, titulo="PREVIEW", limite=500):
        print(f"\n👀 {titulo} (primeiros {limite} caracteres):")
        print("-" * 50)
        print(conteudo[:limite])
        if len(conteudo) > limite:
            print("...")
        print("-" * 50)

    def processar_arquivo(nome_arquivo, codigos_desejados):
        print(f"\n🔥 PROCESSANDO ARQUIVO: {nome_arquivo}")
        print("=" * 80)
        if not analisar_arquivo(nome_arquivo):
            return False
        conteudo, encoding_usado = tentar_ler_arquivo(nome_arquivo)
        if conteudo:
            print(f"\n✅ SUCESSO! Arquivo lido com encoding: {encoding_usado}")
            conteudo_limpo = limpar_certificados_e_assinaturas(conteudo)
            conteudo_filtrado = filtrar_linhas_especificas(conteudo_limpo, codigos_desejados)
            print(f"\n📊 ESTATÍSTICAS FINAIS:")
            print(f"   📏 Tamanho original: {len(conteudo):,} caracteres")
            print(f"   📏 Tamanho limpo: {len(conteudo_limpo):,} caracteres")
            print(f"   📏 Tamanho filtrado: {len(conteudo_filtrado):,} caracteres")
            print(f"   📝 Linhas filtradas: {conteudo_filtrado.count(chr(10)) + 1:,}")
            arquivo_filtrado = salvar_arquivo_limpo(conteudo_filtrado, nome_arquivo, "_FILTRADO")
            if len(conteudo_filtrado.strip()) > 0:
                mostrar_preview(conteudo_filtrado, "CONTEÚDO FILTRADO", 300)
            else:
                print("\n⚠️  Nenhuma linha com os códigos especificados foi encontrada!")
            return True
        else:
            print("\n❌ FALHA: Não foi possível ler o arquivo.")
            return False

    codigos_desejados = ['|0150|', '|0200|', '|C100|', '|C170|', '|D100|', '|D101|', '|D105|', '|D200|', '|D201|', '|D205|']
    print("🚀 LEITOR E FILTRO DE ARQUIVOS SELECIONADOS")
    print("=" * 80)
    if not arquivos:
        print(f"\n❌ Nenhum arquivo selecionado para processar!")
        return
    print(f"\n📋 Encontrados {len(arquivos)} arquivo(s) para processar:")
    for i, arquivo in enumerate(arquivos, 1):
        print(f"   {i}. {os.path.basename(arquivo)}")
    print(f"\n🔄 INICIANDO PROCESSAMENTO...")
    print("=" * 80)
    for i, arquivo in enumerate(arquivos, 1):
        print(f"\n🔢 ARQUIVO {i}/{len(arquivos)}")
        try:
            sucesso = processar_arquivo(arquivo, codigos_desejados)
            if sucesso:
                print(f"✅ Arquivo '{os.path.basename(arquivo)}' processado com sucesso!")
            else:
                print(f"⚠️  Arquivo '{os.path.basename(arquivo)}' com avisos.")
        except Exception as e:
            print(f"❌ Erro ao processar '{os.path.basename(arquivo)}': {str(e)}")
        if i < len(arquivos):
            print("\n" + "🔸" * 80)
    print("\n" + "=" * 80)
    print("📊 RELATÓRIO FINAL DO PROCESSAMENTO")
    print("=" * 80)
    print(f"📁 Total de arquivos processados: {len(arquivos)}")

# Organiza em Json os Dados retirados do Arquivo.txt
def organizar(arquivos):
    def gerar_dict_cfop():
        
        excel_file = pandas.read_excel("tabelas/Tabela_CFOPOperacoesGeradorasCreditos.xls", sheet_name='Tabela I - CFOP x PVA', skiprows=2)

        dados_cfop = {}
        
        for index, row in excel_file.iterrows():
            dados_cfop[row['Código CFOP']] = row['Descrição CFOP']

        return dados_cfop

    dict_cfop = gerar_dict_cfop()

    def gerar_dict_cst():
        excel_file = pandas.read_excel("tabelas/Tabela_CST.xlsx", sheet_name='Planilha1')

        dados_cst = {}
        for index, row in excel_file.iterrows():
            if row['Código'] is not None:
                # Formatando o código para sempre ter 2 dígitos com zero à esquerda
                codigo_formatado = f"{int(row['Código']):02d}"
                dados_cst[codigo_formatado] = [row['Gera Credito/Debito'], row['Tipo de Operação']]
        return dados_cst


    dict_cst = gerar_dict_cst()


    def gera_credito(codigo_cst, cod_cfop=None):
        # Regra especial: CFOP 1556 sempre gera crédito
        if cod_cfop == '1556':
            return True
        if codigo_cst in dict_cst:
            if dict_cst[codigo_cst][0].find('Sim') != -1:
                return True
            else:
                return False
        else:
            # Se o código CST não existe no dicionário, retorna False como padrão
            return False



    for idx, arquivo_path in enumerate(arquivos):
        data_referencia = f'{idx+1:02d}.2024'
        
        # Criar entrada completa para este arquivo
        dados_arquivo = {
                'mes_referencia': data_referencia,
                'dados_notas': [],
                'dados_servicos': [],
                'relatorio_final': []
            }


        with open(arquivo_path, 'r', encoding='utf-8') as arquivo:
            print(f'Processando arquivo {arquivo.name} - {data_referencia}') 
            linhas = arquivo.readlines()
            notas = []  # Lista para armazenar todas as notas
            nota_atual = None
            itens_nota = []  # Lista para armazenar os itens da nota atual
            
            for linha in linhas:
                linha = linha.strip()
                if linha.startswith('|C100|'):
                    # Se já temos uma nota anterior, salva ela com seus itens
                    if nota_atual:
                        notas.append({
                            'nota': nota_atual,
                            'itens': itens_nota
                        })
                    # Inicia nova nota
                    nota_atual = linha
                    itens_nota = []
                elif linha.startswith('|C170|'):
                    # Adiciona item à nota atual
                    itens_nota.append(linha)
                elif nota_atual:  # Outras linhas da nota atual
                    nota_atual += '\n' + linha
            
            # Adiciona a última nota
            if nota_atual:
                notas.append({
                    'nota': nota_atual,
                    'itens': itens_nota
                })
            
            # Criar dicionário estruturado com os dados
            dados_estruturados = []
            relatorio_final = {
                'total_credito_pis': 0,
                'total_credito_cofins': 0,
            }
            
            for i, nota in enumerate(notas):
                # Buscar código do participante na nota
                cod_participante_match = re.search(pattern=r'(FOR\d+)', string=nota['nota'])
                cod_chave_da_nota = re.search(pattern=r'^\|C100\|(?:[^|]*\|){7}([^|]*)', string=nota['nota']).group(1)
                dados_nota = {
                    'id_nota': i + 1,
                    'cod_chave_da_nota': cod_chave_da_nota,
                    'codigo_participante': None,
                    'dados_participante': None,
                    'participante_e_do_estado': None,
                    'valor_total_nota': 0,
                    'valor_total_crédito_pis_nota': 0,
                    'valor_total_crédito_cofins_nota': 0,
                    'dados_itens': [],
                }
                
                if cod_participante_match:
                    cod_participante = cod_participante_match.group(1)
                    dados_nota['codigo_participante'] = cod_participante
                    
                    # Buscar dados do participante no arquivo
                    for linha in linhas:
                        if cod_participante in linha:
                            dados_nota['dados_participante'] = linha.strip()
                            cod_estado_participante = re.search(string=dados_nota['dados_participante'], pattern=r'\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|(\d+)\|').group(1)
                            
                            if cod_estado_participante.startswith('29'):
                                dados_nota['participante_e_do_estado'] = True
                            else:
                                dados_nota['participante_e_do_estado'] = False
                            break
                

                for item in nota['itens']:
                    numero_item = re.search(string=item, pattern=r'^\|C170(?:\|[^|\r\n]*){0}\|([^|\r\n]*?)(?:\||$)').group(1)
                    cod_item_num = re.search(string=item, pattern=r'^\|C170\|[^|]*\|([^|]*)\|').group(1)
                    valor_pis = re.search(string=item, pattern=r'\|([^|]*)\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|$').group(1)
                    valor_cofins = re.search(string=item, pattern=r'\|([^|]+)\|[^|]+\|$').group(1)
                    valor_item = re.search(string=item, pattern=r'^\|C170(?:\|[^|\r\n]*){5}\|([^|\r\n]*?)(?:\||$)').group(1)
                    cod_cfop  = re.search(string=item, pattern=r'^(?:\|[^|]*){10}\|([1-7]\d{3})\|').group(1)
                    cod_cst = re.search(string=item, pattern=r'^\|C170(?:\|[^|\r\n]*){23}\|([^|\r\n]*?)(?:\||$)').group(1)              # Buscar descrição do item no arquivo

                    valor_item_pis_float = round(float(valor_pis.replace(',','.')), 2)
                    valor_item_cofins_float = round(float(valor_cofins.replace(',','.')), 2)
                    valor_item_float = round(float(valor_item.replace(',','.')), 2)


                    dados_nota['valor_total_nota'] = round(dados_nota['valor_total_nota'] + valor_item_float, 2)
                    dados_nota['valor_total_crédito_pis_nota'] = round(dados_nota['valor_total_crédito_pis_nota'] + valor_item_pis_float, 2)
                    dados_nota['valor_total_crédito_cofins_nota'] = round(dados_nota['valor_total_crédito_cofins_nota'] + valor_item_cofins_float, 2)

                    descricao_item = None
                    for linha in linhas:
                        if cod_item_num in linha and '|0200|' in linha:
                            desc_match = re.search(string=linha, pattern=r'\|[^|]*\|[^|]*\|([^|]*)\|')
                            cod_ncm = re.search(string=linha, pattern=r'\|([^|]*)\|[^|]*\|[^|]*\|[^|]*\|[^|]*\|?$').group(1)
                            
                            if desc_match:
                                descricao_item = desc_match.group(1)
                                if cod_ncm.startswith('27'):
                                    combustivel_ou_lubrificante = True
                                else:
                                    combustivel_ou_lubrificante = False
                        
                    
                    
                    item = {
                            'numero_item': numero_item,
                            'codigo_item': cod_item_num,
                            'cod_ncm': cod_ncm,
                            'desc_item': descricao_item,
                            'combustivel_ou_lubrificante': combustivel_ou_lubrificante,
                            'codigo_cfop': cod_cfop,
                            'codigo_cst': cod_cst,
                            'gera_credito': gera_credito(cod_cst, cod_cfop),
                            'tipo': dict_cst[cod_cst][1] if cod_cst in dict_cst else None,
                            'valor_item': valor_item_float,
                            'valor_pis': valor_item_pis_float,
                            'valor_cofins': valor_item_cofins_float,
                    }
                
                    dados_nota['dados_itens'].append(item)
                
                # Adiciona os totais da nota ao relatório final
                relatorio_final['total_credito_pis'] = round(relatorio_final['total_credito_pis'] + dados_nota['valor_total_crédito_pis_nota'], 2)   
                relatorio_final['total_credito_cofins'] = round(relatorio_final['total_credito_cofins'] + dados_nota['valor_total_crédito_cofins_nota'], 2)    
                
                # Adiciona a nota processada à lista
                dados_arquivo['dados_notas'].append(dados_nota)
            

            dados_arquivo['relatorio_final'].append(relatorio_final)

            print(f"Processando arquivo: {arquivo.name}")
            
            # Primeiro, vamos verificar se há linhas D100 no arquivo
            linhas_d100 = [linha for linha in linhas if linha.strip().startswith('|D100|')]
            print(f"Encontradas {len(linhas_d100)} linhas D100 no arquivo")
            
            servicos_tomados = []
            servico_atual = None
            descricao_pis_credito = None
            descricao_cofins_credito = None
            
            for linha in linhas:
                linha = linha.strip()
                if linha.startswith('|D100|'):
                    # Se já temos um serviço anterior, salva ele com seus dados
                    if servico_atual:
                        servicos_tomados.append({
                            'servico': servico_atual,
                            'descricao_pis': descricao_pis_credito,
                            'descricao_cofins': descricao_cofins_credito,
                        })
                    # Inicia novo serviço
                    servico_atual = linha
                    descricao_pis_credito = None
                    descricao_cofins_credito = None
                elif linha.startswith('|D101|'):
                    # Adiciona descrição PIS ao serviço atual
                    descricao_pis_credito = linha
                elif linha.startswith('|D105|'):
                    # Adiciona descrição COFINS ao serviço atual
                    descricao_cofins_credito = linha
                elif servico_atual:  # Outras linhas do serviço atual
                    servico_atual += '\n' + linha

            # Adiciona o último serviço
            if servico_atual:
                servicos_tomados.append({
                    'servico': servico_atual,
                    'descricao_pis': descricao_pis_credito,
                    'descricao_cofins': descricao_cofins_credito,
                })

            print(f"Total de serviços processados: {len(servicos_tomados)}")

            # Inicializar totais dos serviços tomados
            total_credito_pis_servicos_tomados = 0
            total_credito_cofins_servicos_tomados = 0

            dados_servicos = []
            for i, servico in enumerate(servicos_tomados):
                try:
                    cod_servico = re.search(string=servico['servico'], pattern=r'^\|D100(?:\|[^|\r\n]*){8}\|([^|\r\n]*?)(?:\||$)').group(1)
                    valor_servico = re.search(string=servico['servico'], pattern=r'^\|D100(?:\|[^|\r\n]*){13}\|([^|\r\n]*?)(?:\||$)').group(1)
                    
                    # Verificar se as descrições existem antes de tentar fazer regex
                    valor_pis = None
                    if servico['descricao_pis']:
                        pis_match = re.search(string=servico['descricao_pis'], pattern=r'^\|D101(?:\|[^|\r\n]*){6}\|([^|\r\n]*?)(?:\||$)')
                        valor_pis = pis_match.group(1) if pis_match else None
                    
                    valor_cofins = None
                    if servico['descricao_cofins']:
                        cofins_match = re.search(string=servico['descricao_cofins'], pattern=r'^\|D105(?:\|[^|\r\n]*){6}\|([^|\r\n]*?)(?:\||$)')
                        valor_cofins = cofins_match.group(1) if cofins_match else None

                    # Converter valores para float se existirem
                    if valor_pis:
                        valor_servico_pis_float = round(float(valor_pis.replace(',','.')), 2)
                    else:
                        valor_servico_pis_float = 0
                        
                    if valor_cofins:
                        valor_servico_cofins_float = round(float(valor_cofins.replace(',','.')), 2)
                    else:
                        valor_servico_cofins_float = 0

                    dados_servicos_tomados = {
                        'cod_servico': cod_servico,
                        'valor_servico': valor_servico,
                        'valor_pis': valor_servico_pis_float,
                        'valor_cofins': valor_servico_cofins_float,
                        "tipo": "Crédito",
                    }
                    
                    dados_servicos.append(dados_servicos_tomados)
                    print(f"Serviço tomado {i+1}: Código {cod_servico}, Valor PIS {valor_servico_pis_float}, Valor COFINS {valor_servico_cofins_float}")
                    
                    # Acumular totais dos serviços tomados
                    total_credito_pis_servicos_tomados += valor_servico_pis_float
                    total_credito_cofins_servicos_tomados += valor_servico_cofins_float
                    
                    print(f"   -> Total acumulado PIS: R$ {total_credito_pis_servicos_tomados}")
                    print(f"   -> Total acumulado COFINS: R$ {total_credito_cofins_servicos_tomados}")
                    
                except Exception as e:
                    print(f"Erro ao processar serviço tomado {i+1}: {e}")
                    continue
                    
            print("=" * 60)
            print(f"TOTAL FINAL SERVIÇOS TOMADOS:")
            print(f"PIS: R$ {total_credito_pis_servicos_tomados}")
            print(f"COFINS: R$ {total_credito_cofins_servicos_tomados}")
            print("=" * 60)

            # Primeiro, vamos verificar se há linhas D200 no arquivo
            linhas_d200 = [linha for linha in linhas if linha.strip().startswith('|D200|')]
            print(f"Encontradas {len(linhas_d200)} linhas D200 no arquivo")
            
            servicos_prestados = []
            servico_prestado_atual = None
            descricao_pis_debito = None
            descricao_cofins_debito = None
            
            for linha in linhas:
                linha = linha.strip()
                if linha.startswith('|D200|'):
                    # Se já temos um serviço anterior, salva ele com seus dados
                    if servico_prestado_atual:
                        servicos_prestados.append({
                            'servico': servico_prestado_atual,
                            'descricao_pis': descricao_pis_debito,
                            'descricao_cofins': descricao_cofins_debito,
                        })
                    # Inicia novo serviço
                    servico_prestado_atual = linha
                    descricao_pis_debito = None
                    descricao_cofins_debito = None
                elif linha.startswith('|D201|'):
                    # Adiciona descrição PIS ao serviço atual
                    descricao_pis_debito = linha
                elif linha.startswith('|D205|'):
                    # Adiciona descrição COFINS ao serviço atual
                    descricao_cofins_debito = linha
                elif servico_prestado_atual:  # Outras linhas do serviço atual
                    servico_prestado_atual += '\n' + linha

            # Adiciona o último serviço
            if servico_prestado_atual:
                servicos_prestados.append({
                    'servico': servico_prestado_atual,
                    'descricao_pis': descricao_pis_debito,
                    'descricao_cofins': descricao_cofins_debito,
                })

            print(f"Total de serviços processados: {len(servicos_prestados)}")

            # Inicializar totais dos serviços prestados
            total_debito_pis_servicos_prestados = 0
            total_debito_cofins_servicos_prestados = 0

            for i, servico in enumerate(servicos_prestados):
                try:
                    cod_servico = re.search(string=servico['servico'], pattern=r'^\|D200(?:\|[^|\r\n]*){8}\|([^|\r\n]*?)(?:\||$)')
                    cod_servico = cod_servico.group(1) if cod_servico else None
                    
                    valor_servico = re.search(string=servico['servico'], pattern=r'^\|D200(?:\|[^|\r\n]*){13}\|([^|\r\n]*?)(?:\||$)')
                    valor_servico = valor_servico.group(1) if valor_servico else None
                    
                    # Verificar se as descrições existem antes de tentar fazer regex
                    valor_pis = None
                    if servico['descricao_pis']:
                        pis_match = re.search(string=servico['descricao_pis'], pattern=r'^\|D201(?:\|[^|\r\n]*){4}\|([^|\r\n]*?)(?:\||$)')
                        valor_pis = pis_match.group(1) if pis_match else None
                    
                    valor_cofins = None
                    if servico['descricao_cofins']:
                        cofins_match = re.search(string=servico['descricao_cofins'], pattern=r'^\|D205(?:\|[^|\r\n]*){4}\|([^|\r\n]*?)(?:\||$)')
                        valor_cofins = cofins_match.group(1) if cofins_match else None

                    # Converter valores para float se existirem
                    if valor_pis:
                        valor_servico_pis_float = round(float(valor_pis.replace(',','.')), 2)
                    else:
                        valor_servico_pis_float = 0
                        
                    if valor_cofins:
                        valor_servico_cofins_float = round(float(valor_cofins.replace(',','.')), 2)
                    else:
                        valor_servico_cofins_float = 0

                    dados_servicos_prestados = {
                        'cod_servico': cod_servico,
                        'valor_servico': valor_servico,
                        'valor_pis': valor_servico_pis_float,
                        'valor_cofins': valor_servico_cofins_float,
                        "tipo": "Débito",
                    }
                    
                    dados_servicos.append(dados_servicos_prestados)
                    print(f"Serviço prestado {i+1}: Código {cod_servico}, Valor {valor_servico}")
                    
                    # Acumular totais dos serviços prestados
                    total_debito_pis_servicos_prestados += dados_servicos_prestados['valor_pis']
                    total_debito_cofins_servicos_prestados += dados_servicos_prestados['valor_cofins']
                    
                    print(f"   -> Total acumulado débito PIS: R$ {total_debito_pis_servicos_prestados}")
                    print(f"   -> Total acumulado débito COFINS: R$ {total_debito_cofins_servicos_prestados}")
                    
                except Exception as e:
                    print(f"Erro ao processar serviço prestado {i+1}: {e}")
                    continue

            print("=" * 60)
            print(f"TOTAL FINAL SERVIÇOS PRESTADOS:")
            print(f"PIS: R$ {total_debito_pis_servicos_prestados}")
            print(f"COFINS: R$ {total_debito_cofins_servicos_prestados}")
            print("=" * 60)

            dados_arquivo['dados_servicos'].append(dados_servicos)

            # Atualizar relatório final com todos os totais
            relatorio_final_completo = {
                'total_credito_pis_notas': relatorio_final['total_credito_pis'],
                'total_credito_cofins_notas': relatorio_final['total_credito_cofins'],
                'total_credito_pis_servicos_tomados': round(total_credito_pis_servicos_tomados, 2),
                'total_credito_cofins_servicos_tomados': round(total_credito_cofins_servicos_tomados, 2),
                'total_debito_pis_servicos_prestados': round(total_debito_pis_servicos_prestados, 2),
                'total_debito_cofins_servicos_prestados': round(total_debito_cofins_servicos_prestados, 2),
                'total_credito_pis_geral': round(relatorio_final['total_credito_pis'] + total_credito_pis_servicos_tomados, 2),
                'total_credito_cofins_geral': round(relatorio_final['total_credito_cofins'] + total_credito_cofins_servicos_tomados, 2),
                'saldo_pis_final': round((relatorio_final['total_credito_pis'] + total_credito_pis_servicos_tomados) - total_debito_pis_servicos_prestados, 2),
                'saldo_cofins_final': round((relatorio_final['total_credito_cofins'] + total_credito_cofins_servicos_tomados) - total_debito_cofins_servicos_prestados, 2),
            }
            
            dados_arquivo['relatorio_final'] = [relatorio_final_completo]
            
            print(f"=== RELATÓRIO FINAL {data_referencia} ===")
            print(f"Créditos PIS - Notas: R$ {relatorio_final_completo['total_credito_pis_notas']}")
            print(f"Créditos PIS - Serviços Tomados: R$ {relatorio_final_completo['total_credito_pis_servicos_tomados']}")
            print(f"Total Créditos PIS: R$ {relatorio_final_completo['total_credito_pis_geral']}")
            print(f"Débitos PIS - Serviços Prestados: R$ {relatorio_final_completo['total_debito_pis_servicos_prestados']}")
            print(f"Saldo Final PIS: R$ {relatorio_final_completo['saldo_pis_final']}")
            print()
            print(f"Créditos COFINS - Notas: R$ {relatorio_final_completo['total_credito_cofins_notas']}")
            print(f"Créditos COFINS - Serviços Tomados: R$ {relatorio_final_completo['total_credito_cofins_servicos_tomados']}")
            print(f"Total Créditos COFINS: R$ {relatorio_final_completo['total_credito_cofins_geral']}")
            print(f"Débitos COFINS - Serviços Prestados: R$ {relatorio_final_completo['total_debito_cofins_servicos_prestados']}")
            print(f"Saldo Final COFINS: R$ {relatorio_final_completo['saldo_cofins_final']}")
            print("=" * 50)

            # Salvar em JSON para fácil manipulação posterior
            diretorio_json = Path(__file__).parent / 'jsons'
            diretorio_json.mkdir(exist_ok=True)  # Criar diretório se não existir
            
            nome_base = Path(arquivo.name).stem  # Pega o nome do arquivo sem extensão
            caminho_saida = diretorio_json / f'{nome_base}_estruturado.json'
            with open(caminho_saida, 'w', encoding='utf-8') as arquivo_json:
                json.dump(dados_arquivo, arquivo_json, indent=2, ensure_ascii=False)

            print(f"Dados salvos em: {caminho_saida}")

# Verifica a Existência de Erros como [Duplicidade, CFOP, PIS e COFINS zerados e Gera Credito ou não (Dados retirados pelo regex do Json)]
def verificar(arquivos):

    def ler_json_arquivo(nome_arquivo):
        with open(nome_arquivo, 'r', encoding='utf-8') as f:
            return json.load(f)

    def converter_valor(valor_str):
        if isinstance(valor_str, str):
            valor_str = valor_str.replace(",", ".")
        try:
            return float(valor_str)
        except (ValueError, TypeError):
            return 0.0

    def analisar_nota_resumida_com_duplicidade(nota):
        if "dados_itens" not in nota:
            return "", f"Nota {nota.get('id_nota', 'sem id')} está incompleta\n"

        duplicidade_itens = defaultdict(list)
        itens_resultado = []

        for item in nota["dados_itens"]:
            chave = (
                item["desc_item"].strip().lower(),
                converter_valor(item["valor_pis"]),
                converter_valor(item["valor_cofins"])
            )
            duplicidade_itens[chave].append(int(item["numero_item"]))

        grupos_duplicados = [sorted(grupo) for grupo in duplicidade_itens.values() if len(grupo) > 1]
        duplicados = {num for grupo in grupos_duplicados for num in grupo}

        cfop_errados = []
        pis_cofins_zerados = []

        for item in nota["dados_itens"]:
            erros = []
            numero_item = int(item["numero_item"])
            valor_pis = converter_valor(item["valor_pis"])
            valor_cofins = converter_valor(item["valor_cofins"])

            if nota.get("participante_e_do_estado") and not item["codigo_cfop"].startswith("1"):
                erros.append("Erro Cfop")
                cfop_errados.append(numero_item)

            if nota.get("participante_e_do_estado") is False and not item["codigo_cfop"].startswith("2"):
                erros.append("Erro Cfop")
                cfop_errados.append(numero_item)

            if numero_item in duplicados:
                erros.append("Duplicidade")

            if valor_pis == 0.0 and valor_cofins == 0.0:
                erros.append("PIS/COFINS zerado")
                pis_cofins_zerados.append(numero_item)

            status = "Correto" if not erros else " e ".join(erros)
            gera_credito = "Sim" if item.get("gera_credito") else "Não"

            item_texto = (
                f"  Item {numero_item}:\n"
                f"    Status: {status}\n"
                f"    Descrição: {item['desc_item']}\n"
                f"    CFOP: {item['codigo_cfop']}\n"
                f"    CST: {item['codigo_cst']}\n"
                f"    Valor PIS: {valor_pis:.2f}\n"
                f"    Valor COFINS: {valor_cofins:.2f}\n"
                f"    Gera Crédito: {gera_credito}\n"
            )
            itens_resultado.append((numero_item, status, item_texto))

        erros_na_nota = []

        if grupos_duplicados:
            duplicidade_str = ', '.join(f"({', '.join(map(str, grupo))})" for grupo in grupos_duplicados)
            erros_na_nota.append(f"Duplicidade nos Itens: {duplicidade_str}")

        if cfop_errados:
            cfop_str = ', '.join(map(str, sorted(cfop_errados)))
            erros_na_nota.append(f"Erro: CFOP itens {cfop_str}")

        if pis_cofins_zerados:
            piscofins_str = ', '.join(map(str, sorted(pis_cofins_zerados)))
            erros_na_nota.append(f"Erro: PIS/COFINS zerado nos itens {piscofins_str}")

        id_nota = nota.get('id_nota', 'sem id')
        chave_nota = nota.get('cod_chave_da_nota', 'sem chave')

        if not erros_na_nota:
            cabecalho = f"Nota Fiscal: {id_nota} - Correto\nNumero da Nota - {chave_nota}"
            bloco_corretos = cabecalho + "\n" + ''.join(
                texto for _, status, texto in itens_resultado if status == "Correto"
            ) + "-"*40 + "\n"
            return bloco_corretos, ""
        else:
            cabecalho = f"Nota Fiscal: {id_nota} - Erro: " + '; '.join(erros_na_nota)
            cabecalho += f"\nNumero da Nota - {chave_nota}"
            itens_erro = sorted(
                [(num, texto) for num, status, texto in itens_resultado if status != "Correto"],
                key=lambda x: (x[0] not in duplicados, x[0])
            )
            bloco_erros = cabecalho + "\n" + ''.join(texto for _, texto in itens_erro) + "-"*40 + "\n"
            return "", bloco_erros


    def gerar_relatorios_duplos(lista_jsons, pasta_saida='resultado'):
        if not os.path.exists(pasta_saida):
            os.makedirs(pasta_saida)
        for arquivo_json_path, arquivo_origem in zip(lista_jsons, arquivos):
            dados = ler_json_arquivo(arquivo_json_path)
            if not isinstance(dados, dict) or "dados_notas" not in dados or "mes_referencia" not in dados:
                print(f"⚠️ Estrutura inesperada no arquivo: {arquivo_json_path}")
                continue
            nome_arquivo_txt = os.path.splitext(os.path.basename(arquivo_origem))[0]
            nome_base = sanitize_filename(nome_arquivo_txt)
            subpasta = os.path.join(pasta_saida, nome_base)
            try:
                os.makedirs(subpasta, exist_ok=True)
            except Exception as e:
                print(f"Erro ao criar subpasta {subpasta}: {e}")
            relatorio_corretos = f"===== Mês de Referência: {dados['mes_referencia']} =====\n"
            relatorio_erros = f"===== Mês de Referência: {dados['mes_referencia']} =====\n"
            for nota in dados["dados_notas"]:
                bloco_ok, bloco_erro = analisar_nota_resumida_com_duplicidade(nota)
                relatorio_corretos += bloco_ok
                relatorio_erros += bloco_erro
            nome_ok = "relatorio_correto.txt"
            nome_erro = "relatorio_erro.txt"
            nome_final = "relatorio_final.txt"
            nome_servicos = "relatorio_servicos.txt"
            nome_servicos_erro = "relatorio_servicos_erro.txt"
            caminho_ok = os.path.join(subpasta, nome_ok)
            caminho_erro = os.path.join(subpasta, nome_erro)
            caminho_final = os.path.join(subpasta, nome_final)
            caminho_servicos = os.path.join(subpasta, nome_servicos)
            caminho_servicos_erro = os.path.join(subpasta, nome_servicos_erro)
            os.makedirs(os.path.dirname(caminho_ok), exist_ok=True)
            with open(caminho_ok, 'w', encoding='utf-8') as f:
                f.write(relatorio_corretos)
            os.makedirs(os.path.dirname(caminho_erro), exist_ok=True)
            with open(caminho_erro, 'w', encoding='utf-8') as f:
                f.write(relatorio_erros)
            os.makedirs(os.path.dirname(caminho_final), exist_ok=True)
            with open(caminho_final, 'w', encoding='utf-8') as f:
                relatorio_json = dados.get("relatorio_final", [])
                if relatorio_json:
                    for bloco in relatorio_json:
                        for chave, valor in bloco.items():
                            f.write(f"{chave}: {valor}\n")
                else:
                    f.write("Nenhum dado encontrado no campo 'relatorio_final'.\n")

            # Gerar relatório de serviços
            servicos = dados.get('dados_servicos', [])
            relatorio_servicos_corretos = []
            relatorio_servicos_erros = []
            relatorio_servicos_erro = []
            def is_num(val):
                try:
                    float(str(val).replace(',','.'))
                    return True
                except Exception:
                    return False

            nota_idx = 1
            for servico in servicos:
                for s in servico:
                    cod_servico = s.get('cod_servico', '')
                    valor_total = s.get('valor_servico', '')
                    valor_pis = s.get('valor_pis', '')
                    valor_cofins = s.get('valor_cofins', '')
                    tipo = s.get('tipo', '')
                    bloco = []
                    bloco.append(f'Nota - {nota_idx}')
                    bloco.append(f'Codigo de Serviço: "{cod_servico}"')
                    bloco.append(f'Valor Total Serviço: "{valor_total}"')
                    bloco.append(f'Valor Pis: {valor_pis}')
                    bloco.append(f'Valor Cofins: {valor_cofins}')
                    bloco.append(f'Tipo: {tipo}')
                    erros = []
                    erro_num = False
                    if not is_num(valor_total):
                        erros.append('Erro ao converter valor total serviço para número')
                        erro_num = True
                    if not is_num(valor_pis):
                        erros.append('Erro ao converter valor PIS para número')
                        erro_num = True
                    if not is_num(valor_cofins):
                        erros.append('Erro ao converter valor COFINS para número')
                        erro_num = True
                    if not erro_num:
                        v_total = float(str(valor_total).replace(',','.'))
                        v_pis = float(str(valor_pis).replace(',','.'))
                        v_cofins = float(str(valor_cofins).replace(',','.'))
                        if v_total <= 0:
                            erros.append('Valor Total Serviço zerado ou negativo')
                        if v_pis <= 0:
                            erros.append('Valor PIS zerado ou negativo')
                        if v_cofins <= 0:
                            erros.append('Valor COFINS zerado ou negativo')
                    if erros:
                        for erro in erros:
                            bloco.append(f'Erro: {erro}')
                        bloco.append('='*60)
                        relatorio_servicos_erros.append('\n'.join(bloco))
                        # Também adiciona ao relatorio_servicos_erro.txt
                        bloco_erro = []
                        bloco_erro.append(f'Codigo de Serviço: "{cod_servico}"')
                        for erro in erros:
                            bloco_erro.append(f'Erro: {erro}')
                        bloco_erro.append('='*60)
                        relatorio_servicos_erro.append('\n'.join(bloco_erro))
                    else:
                        bloco.append('='*60)
                        relatorio_servicos_corretos.append('\n'.join(bloco))
                    nota_idx += 1
            # Monta o relatório final de serviços com separação
            relatorio_servicos = []
            relatorio_servicos.append('='*60)
            relatorio_servicos.append('Notas Serviços - CORRETOS\n')
            relatorio_servicos.extend(relatorio_servicos_corretos)
            relatorio_servicos.append('='*60)
            relatorio_servicos.append('Notas Serviços - ERROS\n')
            relatorio_servicos.extend(relatorio_servicos_erros)
            # Salvar relatórios de serviços
            with open(caminho_servicos, 'w', encoding='utf-8') as f:
                f.write('\n'.join(relatorio_servicos))
            with open(caminho_servicos_erro, 'w', encoding='utf-8') as f:
                f.write('\n'.join(relatorio_servicos_erro))

            print(f"✅ Relatórios gerados para {arquivo_json_path}:")
            print(f" - {nome_ok}")
            print(f" - {nome_erro}")
            print(f" - {nome_final}")

            # ===== NOVO: Relatórios em Excel =====
            import pandas as pd
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment
            from openpyxl import load_workbook

            # Relatório de Notas Corretas e Erros (com remoção de colunas e centralização)
            corretos = []
            erros = []
            erros_descricoes = []
            for nota in dados["dados_notas"]:
                bloco_ok, bloco_erro = analisar_nota_resumida_com_duplicidade(nota)
                if bloco_ok:
                    corretos.append(nota)
                if bloco_erro:
                    erros.append(nota)
                    erro_descr = bloco_erro.split('\n')[0].replace('Nota Fiscal:', '').replace('-', '').strip()
                    erros_descricoes.append(erro_descr)

            def notas_para_df(lista, add_erro_col=False, erros_descricoes=None):
                linhas = []
                for idx, nota in enumerate(lista):
                    id_nota = nota.get('id_nota')
                    chave = nota.get('cod_chave_da_nota')
                    participante = nota.get('codigo_participante')
                    estado = nota.get('participante_e_do_estado')
                    for item in nota.get('dados_itens', []):
                        linha = {
                            'ID Nota': id_nota,
                            'Chave Nota': chave,
                            'Participante': participante,
                            'Do Estado': estado,
                            'Nº Item': item.get('numero_item'),
                            'NCM': item.get('cod_ncm'),
                            'Descrição Item': item.get('desc_item'),
                            'Combustível/Lubrificante': item.get('combustivel_ou_lubrificante'),
                            'CFOP': item.get('codigo_cfop'),
                            'Gera Crédito': item.get('gera_credito'),
                            'Tipo': item.get('tipo'),
                            'Valor Item': item.get('valor_item'),
                            'Valor PIS': item.get('valor_pis'),
                            'Valor COFINS': item.get('valor_cofins'),
                        }
                        if add_erro_col and erros_descricoes is not None:
                            linha['ERRO'] = erros_descricoes[idx] if idx < len(erros_descricoes) else ''
                        linhas.append(linha)
                return pd.DataFrame(linhas)

            df_corretos = notas_para_df(corretos, add_erro_col=False)
            df_erros = notas_para_df(erros, add_erro_col=True, erros_descricoes=erros_descricoes)

            # Remover colunas indesejadas
            colunas_remover = [
                'Código Item', 'CST', 'Valor Total Nota', 'Valor Total PIS', 'Valor Total COFINS'
            ]
            for col in colunas_remover:
                if col in df_corretos.columns:
                    df_corretos = df_corretos.drop(columns=[col])
                if col in df_erros.columns:
                    df_erros = df_erros.drop(columns=[col])

            nome_ok = "relatorio_correto.xlsx"
            nome_erro = "relatorio_erro.xlsx"
            nome_final = "relatorio_final.xlsx"
            nome_servicos = "relatorio_servicos.xlsx"
            nome_servicos_erro = "relatorio_servicos_erro.xlsx"
            caminho_ok = os.path.join(subpasta, nome_ok)
            caminho_erro = os.path.join(subpasta, nome_erro)
            caminho_final = os.path.join(subpasta, nome_final)
            caminho_servicos = os.path.join(subpasta, nome_servicos)
            caminho_servicos_erro = os.path.join(subpasta, nome_servicos_erro)

            def formatar_excel(path):
                wb = load_workbook(path)
                ws = wb.active
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.column_dimensions[col_letter].width = max_length + 2
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                wb.save(path)

            if not df_corretos.empty:
                df_corretos.to_excel(caminho_ok, index=False)
                formatar_excel(caminho_ok)
            if not df_erros.empty:
                df_erros.to_excel(caminho_erro, index=False)
                formatar_excel(caminho_erro)

            # Relatório Final (totais)
            relatorio_json = dados.get("relatorio_final", [])
            if relatorio_json:
                df_final = pd.DataFrame(relatorio_json)
                # Renomear colunas conforme solicitado
                renomear = {
                    'total_credito_pis_notas': 'Total Credito Pis Notas',
                    'total_credito_cofins_notas': 'Total Credito Cofins Notas',
                    'total_credito_pis_servicos_tomados': 'Total Credito Pis Serviços Tomados',
                    'total_credito_cofins_servicos_tomados': 'Total Credito Cofins Serviços Tomados',
                    'total_debito_pis_servicos_prestados': 'Total Debito Pis Serviços Prestados',
                    'total_debito_cofins_servicos_prestados': 'Total Debito Cofins Serviços Prestados',
                    'total_credito_pis_geral': 'Total Credito Pis Geral',
                    'total_credito_cofins_geral': 'Total Credito Cofins Geral',
                    'saldo_pis_final': 'Saldo Pis Final',
                    'saldo_cofins_final': 'Saldo Cofins Final',
                }
                df_final = df_final.rename(columns=renomear)
                df_final.to_excel(caminho_final, index=False)
                formatar_excel(caminho_final)
            else:
                pd.DataFrame([{"info": "Nenhum dado encontrado no campo 'relatorio_final'."}]).to_excel(caminho_final, index=False)
                formatar_excel(caminho_final)

            # Relatório de Serviços
            servicos = dados.get('dados_servicos', [])
            linhas_servicos = []
            linhas_servicos_erro = []
            def is_num(val):
                try:
                    float(str(val).replace(',','.'))
                    return True
                except Exception:
                    return False
            nota_idx = 1
            for servico in servicos:
                for s in servico:
                    cod_servico = s.get('cod_servico', '')
                    valor_total = s.get('valor_servico', '')
                    valor_pis = s.get('valor_pis', '')
                    valor_cofins = s.get('valor_cofins', '')
                    tipo = s.get('tipo', '')
                    erro_num = False
                    erros = []
                    if not is_num(valor_total):
                        erros.append('Erro ao converter valor total serviço para número')
                        erro_num = True
                    if not is_num(valor_pis):
                        erros.append('Erro ao converter valor PIS para número')
                        erro_num = True
                    if not is_num(valor_cofins):
                        erros.append('Erro ao converter valor COFINS para número')
                        erro_num = True
                    if not erro_num:
                        v_total = float(str(valor_total).replace(',','.'))
                        v_pis = float(str(valor_pis).replace(',','.'))
                        v_cofins = float(str(valor_cofins).replace(',','.'))
                        if v_total <= 0:
                            erros.append('Valor Total Serviço zerado ou negativo')
                        if v_pis <= 0:
                            erros.append('Valor PIS zerado ou negativo')
                        if v_cofins <= 0:
                            erros.append('Valor COFINS zerado ou negativo')
                    linha = {
                        'Nota': nota_idx,
                        'Codigo de Serviço': cod_servico,
                        'Valor Total Serviço': valor_total,
                        'Valor Pis': valor_pis,
                        'Valor Cofins': valor_cofins,
                        'Tipo': tipo
                    }
                    if erros:
                        linha['ERRO'] = '; '.join(erros)
                        linhas_servicos_erro.append(linha)
                    else:
                        linhas_servicos.append(linha)
                    nota_idx += 1
            df_servicos = pd.DataFrame(linhas_servicos)
            df_servicos_erro = pd.DataFrame(linhas_servicos_erro)
            if not df_servicos.empty:
                df_servicos.to_excel(caminho_servicos, index=False)
                formatar_excel(caminho_servicos)
            if not df_servicos_erro.empty:
                df_servicos_erro.to_excel(caminho_servicos_erro, index=False)
                formatar_excel(caminho_servicos_erro)

            print(f"✅ Relatórios Excel gerados para {arquivo_json_path}:")
            print(f" - {nome_ok}")
            print(f" - {nome_erro}")
            print(f" - {nome_final}")

    # Buscar os arquivos JSON gerados na etapa anterior
    diretorio_json = Path(__file__).parent / 'jsons'
    arquivos_jsons_pares = []
    for arquivo in arquivos:
        nome_base = Path(arquivo).stem
        caminho_json = diretorio_json / f'{nome_base}_estruturado.json'
        if caminho_json.exists():
            arquivos_jsons_pares.append((str(caminho_json), arquivo))
    if not arquivos_jsons_pares:
        print("Nenhum arquivo JSON encontrado para verificar!")
        return
    lista_jsons, arquivos_originais = zip(*arquivos_jsons_pares)
    gerar_relatorios_duplos(lista_jsons, pasta_saida='resultado')

# ========== INTERFACE GRÁFICA ==========
style = Style("darkly")
app = style.master
app.title("Análise de SPED")
app.resizable(False, False)

# Tamanho novo + Centralização
largura_janela = 600
altura_janela = 600
centralizar_janela(app, largura_janela, altura_janela)

try:
    app.iconbitmap("assets/icone.ico")
except:
    pass

frame = tk.Frame(app, bg=style.colors.bg)
frame.pack(expand=True)

# Logo
try:
    imagem = Image.open("assets/logo.png")
    imagem = imagem.resize((170, 170))
    imagem_tk = ImageTk.PhotoImage(imagem)
    imagem_label = tk.Label(frame, image=imagem_tk, bg=style.colors.bg)
    imagem_label.image = imagem_tk
    imagem_label.pack(pady=5)
except:
    pass

# Título
titulo = tk.Label(
    frame,
    text="Análise de Speds",
    bg=style.colors.bg,
    fg="white",
    font=("Segoe UI", 16, "bold")
)
titulo.pack(pady=5)

# Botão de seleção de arquivos (múltiplos)
botao_arquivos = Button(
    frame,
    text="Selecionar Arquivo(s) Sped TXT",
    bootstyle="info",
    command=selecionar_arquivos,
    width=35
)
botao_arquivos.pack(pady=5)

# Botão de seleção de pasta
botao_pasta = Button(
    frame,
    text="Selecionar Pasta",
    bootstyle="info",
    command=selecionar_pasta,
    width=35
)
botao_pasta.pack(pady=5)

# Label do arquivo selecionado
label_arquivo = tk.Label(frame, text="Nenhum arquivo selecionado.", bg=style.colors.bg, fg="white")
label_arquivo.pack(pady=5)

# Frame horizontal para botão e temporizador
frame_analise = tk.Frame(frame, bg=style.colors.bg)
frame_analise.pack(pady=10)

botao_executar_analise = Button(
    frame_analise,
    text="Executar Análise",
    bootstyle="primary",
    command=iniciar_analise,
    width=18,
    state="disabled"
)
botao_executar_analise.pack(side="left", padx=(0, 10))

label_temporizador = tk.Label(
    frame_analise,
    text="Tempo: 00:00",
    bg=style.colors.bg,
    fg="white",
    font=("Segoe UI", 10, "bold")
)
label_temporizador.pack(side="left")

# Botão Exportar
botao_exportar = Button(
    frame,
    text="Exportar",
    bootstyle="success",
    command=exportar_resultado,
    width=35,
    state="disabled"
)
botao_exportar.pack(pady=1)

# Animação de carregamento (abaixo do botão Exportar)
gif_label = tk.Label(frame, bg=style.colors.bg)
gif_label.pack(pady=12)

app.mainloop()